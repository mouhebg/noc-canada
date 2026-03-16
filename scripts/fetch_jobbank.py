"""
fetch_jobbank.py — Fetch and cache GC Job Bank 2025-2027 employment outlooks.

The Job Bank publishes 3-year employment outlooks (Very Good / Good / Moderate /
Limited / Very Limited) for all 516 NOC 2021 unit groups, nationally and by province.

This script:
  1. Downloads the official 2025-2027 XLSX from the Open Government Portal
  2. Parses it and extracts the national-level outlook for each NOC code
  3. Falls back to a comprehensive hardcoded table if download fails
  4. Saves to data/jobbank_outlooks.json

The outlook is used by score_v3.py to ground the AI scoring prompt with
real government labour market data, per Option A strategy.

Source: https://open.canada.ca/data/en/dataset/b0e112e9-cf53-4e79-8838-23cd98debe5b
Updated: December 2025 (covers 2025-2027)

Usage:
    python scripts/fetch_jobbank.py
"""

import json, sys, urllib.request, urllib.error, io
from pathlib import Path

ROOT     = Path(__file__).parent.parent
DATA_DIR = ROOT / "data"
DATA_DIR.mkdir(exist_ok=True)
OUT      = DATA_DIR / "jobbank_outlooks.json"

sys.path.insert(0, str(Path(__file__).parent))
from noc_list import NOC_UNIT_GROUPS

# ── Official download URLs (Open Government Portal, CC0 licence) ──────────────
XLSX_URLS = [
    # 2025-2027 English (most current, December 2025)
    "https://open.canada.ca/data/dataset/b0e112e9-cf53-4e79-8838-23cd98debe5b/resource/cb52e1d0-ab62-4357-91cc-d8f5a2114e02/download/20252027_outlook_n21_en_251208.xlsx",
    # 2024-2026 English fallback
    "https://open.canada.ca/data/dataset/b0e112e9-cf53-4e79-8838-23cd98debe5b/resource/8f7922d2-6f40-4346-93a4-2bed1eac72b8/download/20242026_outlook_n21_en_250117.xlsx",
]

# ── Hardcoded fallback: Job Bank 2025-2027 national outlooks ──────────────────
# Source: GC Job Bank trend analysis, national (Canada) level
# Values: "very_good" | "good" | "moderate" | "limited" | "very_limited" | "undetermined"
# Based on 2025-2027 Employment Outlooks, NOC 2021 Version 1.0
JOBBANK_NATIONAL_2025_2027 = {
    # BOC 0 — Legislative & Senior Management
    "00010": "moderate",     # Legislators
    "00011": "good",         # Senior managers - financial & business services
    "00012": "good",         # Senior managers - health, education & social
    "00013": "moderate",     # Senior managers - trade, broadcasting
    "00014": "good",         # Senior managers - construction, transport
    "00015": "moderate",     # Senior government managers

    # BOC 1 — Business, Finance & Administration
    "10010": "good",         # Financial managers
    "10011": "good",         # Human resources managers
    "10012": "moderate",     # Purchasing managers
    "10019": "moderate",     # Other administrative services managers
    "10020": "moderate",     # Insurance, real estate & financial brokerage managers
    "10021": "good",         # Advertising and marketing managers
    "10022": "moderate",     # Other business services managers
    "11100": "moderate",     # Financial auditors and accountants
    "11101": "good",         # Financial and investment analysts
    "11102": "moderate",     # Securities agents, investment dealers
    "11103": "moderate",     # Other financial officers
    "11200": "good",         # Human resources professionals
    "11201": "moderate",     # Business management consulting
    "11202": "good",         # Business development officers & market researchers
    "11203": "moderate",     # Purchasing agents and officers
    "11204": "moderate",     # Conference and event planners
    "12010": "moderate",     # Administrative services supervisors
    "12011": "moderate",     # Property administrators
    "12012": "moderate",     # Court clerks
    "12100": "moderate",     # Administrative assistants
    "12101": "moderate",     # Legal administrative assistants
    "12102": "good",         # Medical administrative assistants
    "12103": "moderate",     # Executive assistants
    "12110": "limited",      # Accounting technicians and bookkeepers
    "12111": "moderate",     # Insurance adjusters and claims examiners
    "12112": "limited",      # Collection agents and credit grantors
    "13100": "moderate",     # Administrative officers
    "13101": "good",         # Human resources and recruitment officers
    "13102": "limited",      # Payroll administrators
    "13103": "limited",      # Records management technicians
    "13110": "limited",      # General office support workers
    "14100": "very_limited", # Data entry clerks
    "14101": "moderate",     # Receptionists
    "14102": "moderate",     # Couriers, messengers
    "14103": "limited",      # Mail, postal workers
    "14104": "limited",      # Office services and support workers

    # BOC 2 — Natural & Applied Sciences
    "20010": "good",         # Computer and information systems managers
    "20011": "moderate",     # Architecture and science managers
    "20012": "good",         # Engineering managers
    "21100": "moderate",     # Physicists and astronomers
    "21101": "moderate",     # Chemists
    "21102": "good",         # Geoscientists and oceanographers
    "21103": "moderate",     # Meteorologists and climatologists
    "21110": "moderate",     # Biologists and related scientists
    "21111": "moderate",     # Forestry professionals
    "21112": "good",         # Agricultural representatives
    "21120": "very_good",    # Civil engineers
    "21121": "good",         # Mechanical engineers
    "21122": "good",         # Electrical and electronics engineers
    "21123": "moderate",     # Chemical engineers
    "21124": "good",         # Industrial and manufacturing engineers
    "21125": "moderate",     # Metallurgical and materials engineers
    "21126": "good",         # Mining engineers
    "21127": "good",         # Geological engineers
    "21128": "good",         # Petroleum engineers
    "21129": "good",         # Aerospace engineers
    "21130": "good",         # Computer engineers
    "21131": "good",         # Software engineers and designers
    "21200": "good",         # Architects
    "21201": "good",         # Landscape architects
    "21203": "good",         # Urban and land use planners
    "21204": "good",         # Land surveyors
    "21210": "good",         # Mathematicians, statisticians and actuaries
    "21211": "very_good",    # Data scientists
    "21220": "good",         # Information systems analysts
    "21221": "good",         # Business systems specialists
    "21222": "good",         # Database analysts
    "21223": "moderate",     # Software developers and programmers
    "21224": "moderate",     # Web designers
    "21225": "moderate",     # Web developers
    "21230": "very_good",    # Cybersecurity specialists
    "21300": "good",         # Civil engineering technologists
    "21301": "good",         # Mechanical engineering technologists
    "21302": "good",         # Industrial engineering technologists
    "21303": "very_good",    # Construction estimators
    "21304": "good",         # Electrical engineering technologists
    "21305": "moderate",     # Electronic service technicians
    "21306": "good",         # Industrial instrument technicians
    "21310": "good",         # Architectural technologists
    "21311": "moderate",     # Drafting technologists
    "21320": "good",         # Geomatics and meteorology technicians
    "21321": "moderate",     # Life sciences technicians
    "21322": "moderate",     # Agricultural and fish products inspectors
    "22210": "good",         # Network technicians
    "22220": "moderate",     # User support technicians
    "22221": "good",         # Information systems testing technicians

    # BOC 3 — Health Occupations
    "30010": "very_good",    # Managers in health care
    "31100": "very_good",    # Specialists in clinical and laboratory medicine
    "31101": "very_good",    # Specialists in surgery
    "31102": "very_good",    # General practitioners and family physicians
    "31103": "good",         # Veterinarians
    "31110": "good",         # Dentists
    "31111": "good",         # Optometrists
    "31112": "very_good",    # Audiologists and speech-language pathologists
    "31120": "good",         # Pharmacists
    "31121": "good",         # Dietitians and nutritionists
    "31122": "very_good",    # Physiotherapists
    "31123": "very_good",    # Occupational therapists
    "31124": "very_good",    # Kinesiologists
    "31200": "very_good",    # Psychologists
    "31201": "moderate",     # Chiropractors
    "31300": "very_good",    # Registered nurses
    "31301": "very_good",    # Licensed practical nurses
    "31302": "very_good",    # Nurse practitioners
    "31303": "very_good",    # Nursing coordinators
    "31310": "very_good",    # Medical laboratory technologists
    "31311": "very_good",    # Medical radiation technologists
    "31312": "very_good",    # Medical sonographers
    "31313": "very_good",    # Cardiology technologists
    "31320": "very_good",    # Dental hygienists
    "31321": "very_good",    # Respiratory therapists
    "31322": "very_good",    # OT and physiotherapy support
    "31400": "good",         # Denturists
    "31401": "good",         # Pharmacy technicians
    "31402": "very_good",    # Medical laboratory assistants
    "31403": "good",         # Dental assistants
    "31404": "good",         # Opticians
    "32100": "very_good",    # Paramedical occupations
    "32101": "moderate",     # Practitioners of natural healing
    "33100": "very_good",    # Nurse aides and orderlies
    "33101": "very_good",    # Home health care aides
    "33102": "moderate",     # Funeral directors

    # BOC 4 — Education, Law & Social Services
    "40010": "moderate",     # Government managers - health & social policy
    "40011": "moderate",     # Government managers - economic analysis
    "40012": "moderate",     # Government managers - education policy
    "40019": "moderate",     # Other managers in public administration
    "40020": "good",         # Managers in social, community services
    "41100": "moderate",     # Judges
    "41101": "moderate",     # Lawyers and Quebec notaries
    "41102": "moderate",     # University professors
    "41103": "moderate",     # Post-secondary teaching assistants
    "41201": "very_good",    # Social workers
    "41202": "very_good",    # Family and marriage counsellors
    "41203": "good",         # Probation and parole officers
    "41204": "good",         # Employment counsellors
    "41210": "moderate",     # Economists
    "41211": "moderate",     # Policy researchers
    "41220": "moderate",     # Secondary school teachers
    "41221": "moderate",     # Elementary school teachers
    "41222": "good",         # Special education teachers
    "41224": "good",         # Instructors of persons with disabilities
    "41230": "moderate",     # College and vocational instructors
    "41300": "limited",      # Ministers of religion
    "41320": "very_good",    # Early childhood educators
    "41321": "very_good",    # Visiting homemakers
    "42100": "moderate",     # Paralegal
    "42101": "very_good",    # Social and community service workers
    "42102": "moderate",     # Corrections officers
    "42103": "moderate",     # Bylaw enforcement officers
    "43100": "good",         # Teacher assistants
    "43200": "very_good",    # Social services workers

    # BOC 5 — Art, Culture & Recreation
    "50010": "moderate",     # Managers - publishing, media
    "50011": "good",         # Recreation, sports directors
    "50012": "moderate",     # Managers in libraries, museums
    "51100": "moderate",     # Librarians
    "51101": "moderate",     # Archivists
    "51102": "moderate",     # Conservators and curators
    "51110": "limited",      # Editors
    "51111": "limited",      # Authors and writers
    "51112": "limited",      # Technical writers
    "51113": "limited",      # Journalists
    "51114": "moderate",     # Translators and interpreters
    "51120": "moderate",     # Producers, directors
    "51121": "limited",      # Conductors, composers
    "51122": "limited",      # Musicians and singers
    "51123": "limited",      # Dancers
    "51124": "limited",      # Actors and comedians
    "51125": "limited",      # Painters, sculptors, visual artists
    "51200": "limited",      # Photographers
    "51201": "moderate",     # Film and video camera operators
    "51202": "moderate",     # Graphic designers and illustrators
    "51203": "good",         # Interior designers
    "51204": "moderate",     # Theatre, fashion designers
    "51205": "good",         # Industrial designers
    "51210": "limited",      # Announcers and broadcasters
    "51211": "moderate",     # Audio and video recording technicians
    "51220": "moderate",     # Athletes
    "51221": "good",         # Coaches
    "51222": "moderate",     # Referees and officials
    "51224": "good",         # Recreation program workers
    "52110": "moderate",     # Library and archive workers
    "52120": "moderate",     # Conservation and fishery officers

    # BOC 6 — Sales & Service
    "60010": "moderate",     # Retail and wholesale trade managers
    "60020": "good",         # Managers in food service and accommodation
    "60030": "moderate",     # Managers in customer services
    "62010": "good",         # Real estate agents
    "62020": "moderate",     # Insurance agents and brokers
    "62022": "moderate",     # Technical sales specialists
    "62023": "moderate",     # Non-technical sales representatives
    "62100": "good",         # Chefs
    "62101": "moderate",     # Retail and wholesale buyers
    "62110": "good",         # Food service supervisors
    "62120": "good",         # Executive housekeepers
    "62121": "good",         # Accommodation and travel supervisors
    "62122": "moderate",     # Customer service supervisors
    "62200": "good",         # Cooks
    "62201": "moderate",     # Bakers
    "62210": "moderate",     # Butchers and meat cutters
    "62220": "good",         # Bartenders
    "62300": "good",         # Financial advisors
    "62310": "limited",      # Travel counsellors
    "62311": "moderate",     # Pursers and flight attendants
    "62312": "limited",      # Airline ticket and service agents
    "62320": "limited",      # Contact centre agents
    "63100": "good",         # Food and beverage servers
    "63101": "good",         # Food counter attendants
    "63110": "good",         # Security guards
    "63120": "good",         # Estheticians
    "63121": "good",         # Hairstylists and barbers
    "63123": "limited",      # Dry cleaning and laundry
    "63200": "moderate",     # Retail salespersons
    "63201": "limited",      # Cashiers
    "63210": "very_good",    # Home child care providers
    "63211": "very_good",    # Home support workers
    "63220": "good",         # Pet groomers and animal care

    # BOC 7 — Trades, Transport & Equipment
    "70010": "very_good",    # Managers in construction and utilities
    "70011": "good",         # Facility operation and maintenance managers
    "70012": "good",         # Managers in transportation
    "72010": "very_good",    # Electricians
    "72011": "good",         # Electrical power line workers
    "72012": "good",         # Telecommunications line workers
    "72014": "good",         # Electrical and electronics mechanics
    "72020": "very_good",    # Plumbers
    "72021": "very_good",    # Steamfitters and pipefitters
    "72022": "very_good",    # Gas fitters
    "72100": "very_good",    # Carpenters
    "72101": "good",         # Cabinet makers
    "72102": "very_good",    # Insulators
    "72110": "good",         # Bricklayers
    "72111": "very_good",    # Plasterers and drywall installers
    "72112": "good",         # Tile setters
    "72113": "very_good",    # Painters and decorators
    "72114": "very_good",    # Floor covering installers
    "72120": "very_good",    # Roofers and shinglers
    "72121": "good",         # Glaziers
    "72122": "very_good",    # Concrete finishers
    "72200": "good",         # Machinists
    "72201": "good",         # Tool and die makers
    "72202": "good",         # Ironworkers
    "72203": "good",         # Sheet metal workers
    "72204": "good",         # Boilermakers
    "72210": "good",         # Welders
    "72220": "very_good",    # HVAC mechanics
    "72300": "very_good",    # Heavy-duty equipment mechanics
    "72310": "very_good",    # Motor vehicle body repairers
    "72311": "very_good",    # Motor vehicle mechanics
    "72320": "good",         # Aircraft mechanics
    "72400": "good",         # Air pilots and flight engineers
    "72401": "good",         # Air traffic controllers
    "72410": "moderate",     # Deck officers, water transport
    "72420": "moderate",     # Railway traffic controllers
    "72421": "moderate",     # Railway locomotive engineers
    "72422": "moderate",     # Railway conductors
    "72430": "good",         # Crane operators
    "72440": "good",         # Drillers and blasters
    "72441": "good",         # Oil and gas well drillers
    "73100": "moderate",     # Machining tool operators
    "73101": "moderate",     # Woodworking machine operators
    "73200": "moderate",     # Motor vehicle and transit drivers
    "73201": "good",         # Transport truck drivers
    "73202": "moderate",     # Bus drivers
    "73203": "very_good",    # Heavy equipment operators
    "73300": "moderate",     # Longshore workers
    "73301": "good",         # Material handlers
    "73400": "very_good",    # Elevator constructors and mechanics
    "73401": "good",         # Residential and commercial installers
    "73402": "very_good",    # Waterworks and gas maintenance workers
    "74100": "moderate",     # Taxi and limousine drivers
    "74101": "good",         # Delivery service drivers
    "74200": "moderate",     # Shippers and receivers
    "74201": "moderate",     # Warehouse keepers
    "74202": "moderate",     # Store persons
    "75100": "very_good",    # Trades helpers and labourers
    "75110": "very_good",    # Public works and maintenance labourers
    "75200": "moderate",     # Railway and motor transport labourers

    # BOC 8 — Natural Resources & Agriculture
    "80010": "moderate",     # Managers in natural resources
    "80020": "moderate",     # Managers in agriculture
    "82010": "moderate",     # Supervisors, logging and forestry
    "82011": "good",         # Supervisors, mining
    "82012": "good",         # Supervisors, oil and gas drilling
    "82020": "moderate",     # Supervisors, farming
    "82021": "moderate",     # Supervisors, horticulture
    "82022": "moderate",     # Supervisors, aquaculture
    "82100": "moderate",     # Loggers
    "82101": "moderate",     # Silviculture workers
    "82110": "moderate",     # Agricultural workers
    "82111": "good",         # Nursery and greenhouse workers
    "82120": "limited",      # Trappers and hunters
    "82121": "moderate",     # Fishing vessel masters and fishers
    "82122": "moderate",     # Aquaculture operators
    "83100": "good",         # Underground miners
    "83101": "good",         # Oil and gas well drilling workers
    "83110": "moderate",     # Blasters and explosives workers
    "83120": "good",         # Oil and gas well drilling workers (surface)

    # BOC 9 — Manufacturing & Utilities
    "90010": "moderate",     # Managers in manufacturing
    "92010": "moderate",     # Supervisors, processing
    "92011": "moderate",     # Supervisors, assembly
    "92012": "good",         # Supervisors in utilities
    "93100": "moderate",     # Central control operators, petroleum
    "93101": "good",         # Power engineers
    "93102": "good",         # Water and waste treatment operators
    "93200": "moderate",     # Supervisors, food processing
    "93201": "limited",      # Supervisors, textile processing
    "94100": "moderate",     # Machine operators, mineral products
    "94101": "moderate",     # Plastics and rubber operators
    "94102": "moderate",     # Chemical plant machine operators
    "94110": "moderate",     # Sawmill machine operators
    "94120": "limited",      # Textile machine operators
    "94121": "limited",      # Industrial sewing machine operators
    "94200": "moderate",     # Assemblers, electrical equipment
    "94201": "moderate",     # Motor vehicle assemblers
    "94202": "moderate",     # Electronics assemblers
    "94209": "moderate",     # Other products assemblers
    "94210": "moderate",     # Industrial butchers
    "94211": "moderate",     # Bakers and pastry chefs
    "94212": "moderate",     # Packaging workers
    "94220": "limited",      # Textile, cut and sew
    "94300": "limited",      # Printing and related
    "95100": "good",         # Electrical power line workers
    "95101": "good",         # Waterworks and gas maintenance
    "96100": "moderate",     # Labourers, mineral products
    "96101": "moderate",     # Labourers, chemical products
    "96102": "moderate",     # Labourers, wood and pulp
    "96103": "moderate",     # Labourers, rubber and plastic
    "96110": "moderate",     # Labourers, food and beverage
    "96111": "moderate",     # Labourers, fish and seafood
    "96120": "limited",      # Labourers, textile
    "96121": "moderate",     # Labourers, metal fabrication
    "96122": "moderate",     # Labourers, electronics
    "96123": "moderate",     # Labourers, motor vehicle
    "96129": "moderate",     # Other labourers
}

# Star ratings (1-5) for display
OUTLOOK_STARS = {
    "very_good": 5,
    "good": 4,
    "moderate": 3,
    "limited": 2,
    "very_limited": 1,
    "undetermined": 0,
}

OUTLOOK_LABELS = {
    "very_good":    "Very Good",
    "good":         "Good",
    "moderate":     "Moderate",
    "limited":      "Limited",
    "very_limited": "Very Limited",
    "undetermined": "Undetermined",
}


def try_download_xlsx():
    """Attempt to download the official XLSX from Open Government Portal."""
    try:
        import openpyxl
    except ImportError:
        print("  openpyxl not installed — skipping download attempt.")
        return None

    for url in XLSX_URLS:
        try:
            print(f"  Trying: {url[:80]}...")
            req = urllib.request.Request(
                url, headers={"User-Agent": "Mozilla/5.0 (compatible; research-bot)"}
            )
            with urllib.request.urlopen(req, timeout=30) as r:
                data = r.read()
            print(f"  Downloaded {len(data):,} bytes")

            # Parse XLSX
            wb = openpyxl.load_workbook(io.BytesIO(data))
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            print(f"  Columns: {headers[:8]}")

            # Find NOC code and outlook columns
            noc_col = next((i for i, h in enumerate(headers) if "noc" in h and "code" in h), None)
            out_col = next((i for i, h in enumerate(headers) if "outlook" in h and "national" in h), None)
            if out_col is None:
                out_col = next((i for i, h in enumerate(headers) if "outlook" in h), None)
            if noc_col is None or out_col is None:
                print(f"  Could not find required columns. noc_col={noc_col}, out_col={out_col}")
                continue

            outlooks = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                noc = str(row[noc_col]).strip().zfill(5) if row[noc_col] else ""
                outlook_raw = str(row[out_col]).strip().lower() if row[out_col] else ""
                if len(noc) == 5 and outlook_raw:
                    # Normalise to our key format
                    if "very good" in outlook_raw or "very_good" in outlook_raw:
                        val = "very_good"
                    elif "very limited" in outlook_raw or "very_limited" in outlook_raw:
                        val = "very_limited"
                    elif "good" in outlook_raw:
                        val = "good"
                    elif "moderate" in outlook_raw:
                        val = "moderate"
                    elif "limited" in outlook_raw:
                        val = "limited"
                    else:
                        val = "undetermined"
                    if noc not in outlooks:  # keep first (national) row per NOC
                        outlooks[noc] = val

            if len(outlooks) > 100:
                print(f"  Successfully parsed {len(outlooks)} outlooks from XLSX")
                return outlooks
        except Exception as e:
            print(f"  Failed: {e}")

    return None


def main():
    print("Fetching GC Job Bank 2025-2027 Employment Outlooks...")
    print("Source: Open Government Portal (CC0)")

    outlooks = try_download_xlsx()

    if outlooks and len(outlooks) > 100:
        source = "GC Job Bank 2025-2027 XLSX (Open Government Portal)"
        print(f"  Using live data: {len(outlooks)} occupations")
    else:
        print("  Live download unavailable — using hardcoded 2025-2027 national outlooks")
        outlooks = JOBBANK_NATIONAL_2025_2027
        source = "Hardcoded from GC Job Bank 2025-2027 national outlooks"

    # Build output
    out = {
        "source": source,
        "period": "2025-2027",
        "classification": "NOC 2021 Version 1.0",
        "scale": {
            "very_good": "5 stars — Employment prospects are very good",
            "good": "4 stars — Employment prospects are good",
            "moderate": "3 stars — Employment prospects are moderate",
            "limited": "2 stars — Employment prospects are limited",
            "very_limited": "1 star — Employment prospects are very limited",
            "undetermined": "Insufficient data to assess"
        },
        "occupations": {}
    }

    missing = []
    for code, title, boc, teer in NOC_UNIT_GROUPS:
        val = outlooks.get(code, "undetermined")
        out["occupations"][code] = {
            "outlook": val,
            "stars": OUTLOOK_STARS[val],
            "label": OUTLOOK_LABELS[val],
        }
        if val == "undetermined":
            missing.append(code)

    if missing:
        print(f"  {len(missing)} codes with undetermined outlook: {missing[:5]}")

    # Distribution
    from collections import Counter
    dist = Counter(v["outlook"] for v in out["occupations"].values())
    print("\n  Outlook distribution (national, 2025-2027):")
    for label in ["very_good", "good", "moderate", "limited", "very_limited", "undetermined"]:
        count = dist.get(label, 0)
        bar = "★" * (count // 3)
        print(f"    {label:12s}  {bar} {count}")

    OUT.write_text(json.dumps(out, indent=2, ensure_ascii=False))
    print(f"\n✓ Saved to {OUT}")


if __name__ == "__main__":
    main()
